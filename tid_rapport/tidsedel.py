import openpyxl
from openpyxl.styles import Border, PatternFill, Font, Alignment, colors, Side
from .settingsexl import (
    colum_width_settings,
    row_heigt_settings,
    merge_cells_settings,
    font_settings,
    set_initial_text,
    font_normal_settings,
    border_font_normal,
    formula_tot,
    summa_tot,
)

from .models import Tid


def width_as_mm(mm):
    result = mm * 0.4048583
    return result


def height_as_mm(mm):
    result = mm * 2.83286119
    return result


def skapa_tidsedel(ar, start, stopp, response, user):

    queryset = Tid.objects.filter(vecka__range=(start, stopp)).filter(ar=ar).filter(user_id=user)

    wb = openpyxl.Workbook()
    sheet = wb.active
    font_logga = Font(name="Arial", size=14, bold=True)
    font_tiduppgift = Font(name="Arial", size=16, bold=True)
    font_small = Font(name="Arial", size=8, bold=True)
    font_head = Font(name="Arial", size=11, bold=True)
    font_normal = Font(name="Arial", size=11, bold=False)
    align_tiduppgift = Alignment(horizontal="center", vertical="center")
    align_center = Alignment(horizontal="center")
    align_right = Alignment(horizontal="right")
    align_left = Alignment(horizontal="left")
    fill_yellow = PatternFill(
        patternType="solid", start_color=colors.YELLOW, end_color=colors.YELLOW
    )
    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

    #  Bredd kolumer, höjd på rader, Sammanfogning av celler.
    for key in colum_width_settings:
        sheet.column_dimensions[key].width = width_as_mm(colum_width_settings[key])

    for key in row_heigt_settings:
        sheet.row_dimensions[int(key)].height = height_as_mm(row_heigt_settings[key])

    for item in merge_cells_settings:
        sheet.merge_cells(item)

    # Text och kanter
    for key in font_settings:
        c = sheet[key]
        if font_settings[key] == "font_logga":
            c.font = font_logga
            c.fill = fill_yellow
        elif font_settings[key] == "font_tiduppgift":
            c.font = font_tiduppgift
            c.alignment = align_tiduppgift
            c.border = border_all
        elif font_settings[key] == "font_small":
            c.font = font_small
            c.border = border_all
        elif font_settings[key] == "font_head":
            c.font = font_head
            c.alignment = align_center
            c.border = border_all

    sheet["A21"].font = font_normal
    sheet["A21"].alignment = align_right
    sheet["A21"].border = border_all

    for item in border_font_normal:
        sheet[item].font = font_normal
        sheet[item].border = border_all
        sheet[item].alignment = align_left

    for item in font_normal_settings:
        i = 8
        while i < 21:
            sheet[item + str(i)].font = font_normal
            sheet[item + str(i)].border = border_all
            i += 1

    #  Start text
    for key in set_initial_text:
        sheet[key] = set_initial_text[key]

    #  Formler
    for item in formula_tot:
        i = 8
        while i < 21:
            sheet[item + str(i)] = "=SUM(D" + str(i) + ":K" + str(i) + ")"
            i += 1

    for key in summa_tot:
        sheet[key] = summa_tot[key]

    i = 8
    while i < 21:
        sheet["N" + str(i)] = "=IF($L" + str(i) + ">40, $L" + str(i) + "-40, 0)"
        i += 1

    i = 8
    while i < 21:
        sheet["O" + str(i)] = "=SUM(I" + str(i) + ":J" + str(i) + ")"
        i += 1

    #  Ifyllning av tidsedel.
    querystart = Tid.objects.filter(vecka__exact=start).filter(ar=ar).filter(user_id=user)
    querystopp = Tid.objects.filter(vecka__exact=stopp).filter(ar=ar).filter(user_id=user)

    for query in querystart:
        sheet['K2'] = query.ar
        sheet['A4'] = query.get_full_name()
        sheet['K4'] = str(query.projektnr.anlaggning)
        sheet['H6'] = query.projektnr.ort
        sheet['A6'] = query.projektnr.kund
        startvecka = str(query.vecka) + '-'
        sheet['O2'] = query.vecka

    for query in querystopp:
        if start != stopp:
            sheet['O2'] = startvecka + str(query.vecka)
    i = 8
    for query in queryset:
        sheet['A' + str(i)] = query.vecka
        sheet['B' + str(i)] = query.projektnr.projektnr
        sheet['C' + str(i)] = query.projektnr.anlaggning + ', ' + query.projektnr.ort
        sheet['D' + str(i)] = query.mon
        sheet['E' + str(i)] = query.tis
        sheet['F' + str(i)] = query.ons
        sheet['G' + str(i)] = query.tors
        sheet['H' + str(i)] = query.fre
        sheet['I' + str(i)] = query.lor
        sheet['J' + str(i)] = query.son
        sheet['K' + str(i)] = query.restid
        sheet['P' + str(i)] = query.trakt
        i += 1

    wb.save(filename=response)
    return response
